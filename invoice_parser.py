import pandas as pd
import openpyxl
import re
from pathlib import Path
from datetime import datetime
import os
import sys
import json

def get_current_month():
    """Возвращает текущий месяц на русском"""
    month_names = {
        '01': 'январь', '02': 'февраль', '03': 'март', '04': 'апрель',
        '05': 'май', '06': 'июнь', '07': 'июль', '08': 'август',
        '09': 'сентябрь', '10': 'октябрь', '11': 'ноябрь', '12': 'декабрь'
    }
    
    current_month = datetime.now().month
    month_name = month_names.get(str(current_month).zfill(2))
    
    print(f"📅 Используется месяц: {month_name}", file=sys.stderr)
    return month_name

def find_table_structure(ws):
    """Находит структуру таблицы по ключевым заголовкам - ТОЧНАЯ КОПИЯ ИЗ ТВОЕГО СКРИПТА"""
    print("🔍 Ищу структуру таблицы...", file=sys.stderr)
    
    headers_positions = {}
    
    for row in ws.iter_rows():
        for cell in row:
            if cell.value:
                cell_value = str(cell.value).strip()
                
                if "Товары (работы, услуги)" in cell_value:
                    headers_positions['description'] = (cell.row, cell.column)
                    print(f"📋 Найден заголовок 'Товары' в строке {cell.row}, столбце {cell.column}", file=sys.stderr)
                elif "Сумма" in cell_value and cell_value != "Сумма с НДС":
                    headers_positions['amount'] = (cell.row, cell.column)
                    print(f"💰 Найден заголовок 'Сумма' в строке {cell.row}, столбце {cell.column}", file=sys.stderr)
                elif "№" == cell_value and cell.column < 10:
                    headers_positions['number'] = (cell.row, cell.column)
                    print(f"🔢 Найден заголовок '№' в строке {cell.row}, столбце {cell.column}", file=sys.stderr)
                elif "Кол-во" in cell_value:
                    headers_positions['quantity'] = (cell.row, cell.column)
                    print(f"📦 Найден заголовок 'Кол-во' в строке {cell.row}, столбце {cell.column}", file=sys.stderr)
                elif "Ед." in cell_value:
                    headers_positions['unit'] = (cell.row, cell.column)
                    print(f"📏 Найден заголовок 'Ед.' в строке {cell.row}, столбце {cell.column}", file=sys.stderr)
                elif "Цена" in cell_value:
                    headers_positions['price'] = (cell.row, cell.column)
                    print(f"🏷️ Найден заголовок 'Цена' в строке {cell.row}, столбце {cell.column}", file=sys.stderr)
    
    return headers_positions

def extract_data_from_description(description):
    """Извлекает дату, маршрут, гос. номер и фамилию водителя из описания - ТОЧНАЯ КОПИЯ"""
    description_str = str(description)
    
    # Маршрут (все до первой запятой)
    route = description_str.split(',')[0].strip()
    
    # Дата из текста (формат "от 06.09.25")
    date_match = re.search(r'от\s+(\d{2}\.\d{2}\.\d{2})', description_str)
    date_str = date_match.group(1) if date_match else "Дата не найдена"
    
    # Гос. номер - ищем 3 цифры подряд
    plate_match = re.search(r'(\d{3})', description_str)
    car_plate = plate_match.group(1) if plate_match else "Неизвестно"
    
    # Фамилия водителя
    driver_match = re.search(r',\s*([А-Я][а-я]+)\s+[А-Я]\.[А-Я]\.', description_str)
    if driver_match:
        driver_name = driver_match.group(1)
    else:
        alt_driver_match = re.search(r',\s*([А-Я][а-я]+)', description_str)
        driver_name = alt_driver_match.group(1) if alt_driver_match else "Фамилия не найдена"
    
    return route, date_str, car_plate, driver_name

def parse_invoice_file(file_path):
    """Парсит один файл счета и возвращает данные - ТОЧНАЯ КОПИЯ ИЗ ТВОЕГО СКРИПТА"""
    try:
        print(f"\n🔍 Обрабатываю файл: {file_path.name}", file=sys.stderr)
        wb = openpyxl.load_workbook(file_path, data_only=True)
        ws = wb.active
        
        headers = find_table_structure(ws)
        
        if not headers.get('description') or not headers.get('amount'):
            print("⚠️ Не найдена полная структура таблицы", file=sys.stderr)
            return []
        
        header_row = max(h[0] for h in headers.values())
        description_col = headers['description'][1]
        amount_col = headers['amount'][1]
        
        print(f"📊 Структура таблицы определена:", file=sys.stderr)
        print(f"   - Строка данных: {header_row + 1}", file=sys.stderr)
        print(f"   - Столбец описания: {description_col}", file=sys.stderr)
        print(f"   - Столбец суммы: {amount_col}", file=sys.stderr)
        
        parsed_data = []
        row_num = header_row + 1
        processed_count = 0
        current_empty_rows = 0
        max_empty_rows = 5
        
        print(f"🔎 Читаю данные таблицы...", file=sys.stderr)
        
        while current_empty_rows < max_empty_rows:
            description_cell = ws.cell(row=row_num, column=description_col)
            description = description_cell.value
            
            if not description:
                current_empty_rows += 1
                row_num += 1
                continue
                
            current_empty_rows = 0
            description_str = str(description)
            
            if any(word in description_str.lower() for word in ['итого', 'всего', 'итог', 'сумма']):
                row_num += 1
                continue
            
            amount_cell = ws.cell(row=row_num, column=amount_col)
            amount = amount_cell.value
            
            if amount is not None:
                try:
                    if isinstance(amount, str) and any(char.isalpha() for char in amount.replace(' ', '').replace(',', '.')):
                        row_num += 1
                        continue
                    
                    amount_str = str(amount).replace(' ', '').replace(',', '.')
                    amount_value = float(amount_str)
                    
                    route, date_str, car_plate, driver_name = extract_data_from_description(description_str)
                    
                    if car_plate != "Неизвестно" and amount_value > 0:
                        parsed_data.append({
                            'Дата': date_str,
                            'Маршрут': route,
                            'Стоимость': amount_value,
                            'Гос_номер': car_plate,
                            'Водитель': driver_name,
                            'Источник': file_path.name,
                            'Строка': row_num
                        })
                        processed_count += 1
                        print(f"✅ Строка {row_num}: {date_str} | {route[:20]}... | {car_plate} | {driver_name} | {amount_value:,.0f} руб.", file=sys.stderr)
                    
                except (ValueError, TypeError):
                    pass
            
            row_num += 1
            
            if row_num > header_row + 1000:
                print("⚠️ Достигнуто ограничение в 1000 строк", file=sys.stderr)
                break
        
        print(f"📊 Обработано записей: {processed_count}", file=sys.stderr)
        return parsed_data
        
    except Exception as e:
        print(f"❌ Ошибка при обработке {file_path}: {e}", file=sys.stderr)
        return []

def main():
    """Основная функция для веб-интеграции"""
    try:
        print("🚀 Запуск Python парсера...", file=sys.stderr)
        
        # Обрабатываем файлы из аргументов командной строки
        file_paths = []
        if len(sys.argv) > 1:
            file_paths = [Path(f) for f in sys.argv[1:]]
        else:
            # Для обратной совместимости
            input_folder = Path("uploads")
            if input_folder.exists():
                file_paths = list(input_folder.glob("*.xlsx"))
        
        print(f"📁 Найдено файлов: {len(file_paths)}", file=sys.stderr)
        
        if not file_paths:
            error_result = {
                "success": False,
                "error": "Нет Excel-файлов для обработки",
                "data": []
            }
            print(json.dumps(error_result, ensure_ascii=False))
            return
        
        # Собираем все данные
        all_data = []
        
        for file_path in file_paths:
            if '~' in file_path.name:
                continue
            file_data = parse_invoice_file(file_path)
            all_data.extend(file_data)
        
        if not all_data:
            error_result = {
                "success": False,
                "error": "Не найдено данных для обработки", 
                "data": []
            }
            print(json.dumps(error_result, ensure_ascii=False))
            return
        
        # Создаем DataFrame
        df = pd.DataFrame(all_data)
        
        # Статистика
        total_amount = df['Стоимость'].sum()
        unique_cars = [x for x in df['Гос_номер'].unique() if x != "Неизвестно"]
        unique_drivers = [x for x in df['Водитель'].unique() if x != "Фамилия не найдена"]
        
        # Результат для веба
        result = {
            "success": True,
            "message": f"Обработано {len(all_data)} записей",
            "statistics": {
                "total_records": len(all_data),
                "total_amount": total_amount,
                "unique_cars": len(unique_cars),
                "unique_drivers": len(unique_drivers),
                "car_list": unique_cars,
                "driver_list": unique_drivers
            },
            "data": all_data
        }
        
        print(json.dumps(result, ensure_ascii=False))
        
    except Exception as e:
        error_result = {
            "success": False,
            "error": f"Ошибка выполнения: {str(e)}",
            "data": []
        }
        print(json.dumps(error_result, ensure_ascii=False))

if __name__ == "__main__":
    main()
